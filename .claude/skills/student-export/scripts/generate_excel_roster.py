#!/usr/bin/env python3
"""
Generate Excel roster export for Guitar CRM students.

Usage:
    python generate_excel_roster.py --data roster.json --output roster.xlsx

Or import and use programmatically:
    from generate_excel_roster import create_roster_export
    create_roster_export(students_data, output_path)
"""

import json
import argparse
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Status colors (ARGB format for openpyxl)
STATUS_FILLS = {
    'to_learn': PatternFill(start_color='FFFFEB9C', end_color='FFFFEB9C', fill_type='solid'),
    'started': PatternFill(start_color='FFB4C6E7', end_color='FFB4C6E7', fill_type='solid'),
    'remembered': PatternFill(start_color='FFE2D5F4', end_color='FFE2D5F4', fill_type='solid'),
    'mastered': PatternFill(start_color='FFC6EFCE', end_color='FFC6EFCE', fill_type='solid'),
}

HEADER_FILL = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFFFF')
THIN_BORDER = Border(
    left=Side(style='thin', color='FFDEE2E6'),
    right=Side(style='thin', color='FFDEE2E6'),
    top=Side(style='thin', color='FFDEE2E6'),
    bottom=Side(style='thin', color='FFDEE2E6')
)


def create_roster_export(data: list, output_path: str) -> str:
    """
    Create an Excel roster export for all students.

    Args:
        data: List of student records, each containing:
            - user: {firstName, lastName, email, isActive, created_at}
            - lessonCount: int
            - lessonsThisMonth: int
            - songCounts: {to_learn, started, remembered, mastered}
            - lastLessonDate: string or None
        output_path: Where to save the Excel file

    Returns:
        Path to the generated Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Roster"

    # Define columns
    columns = [
        ('Name', 25),
        ('Email', 30),
        ('Status', 10),
        ('Total Lessons', 14),
        ('This Month', 12),
        ('To Learn', 10),
        ('Started', 10),
        ('Remembered', 12),
        ('Mastered', 10),
        ('Last Lesson', 14),
        ('Member Since', 14),
    ]

    # Set up headers
    for col, (header, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Add data rows
    for row_num, student in enumerate(data, start=2):
        user = student.get('user', {})
        song_counts = student.get('songCounts', {})

        # Build name
        name = f"{user.get('firstName', '')} {user.get('lastName', '')}".strip()
        if not name:
            name = user.get('email', 'Unknown').split('@')[0]

        # Format dates
        last_lesson = student.get('lastLessonDate')
        if last_lesson:
            try:
                last_lesson = datetime.fromisoformat(last_lesson.replace('Z', '+00:00')).strftime('%Y-%m-%d')
            except:
                last_lesson = str(last_lesson)[:10]
        else:
            last_lesson = 'Never'

        member_since = user.get('created_at', '')
        if member_since:
            try:
                member_since = datetime.fromisoformat(member_since.replace('Z', '+00:00')).strftime('%Y-%m-%d')
            except:
                member_since = str(member_since)[:10]

        # Row data
        row_data = [
            name,
            user.get('email', ''),
            'Active' if user.get('isActive', True) else 'Inactive',
            student.get('lessonCount', 0),
            student.get('lessonsThisMonth', 0),
            song_counts.get('to_learn', 0),
            song_counts.get('started', 0),
            song_counts.get('remembered', 0),
            song_counts.get('mastered', 0),
            last_lesson,
            member_since,
        ]

        for col, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical='center')

            # Center align numeric columns
            if col >= 4 and col <= 9:
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Highlight mastered count if >= 5
            if col == 9 and isinstance(value, int) and value >= 5:
                cell.fill = STATUS_FILLS['mastered']

            # Highlight inactive students
            if col == 3 and value == 'Inactive':
                cell.fill = PatternFill(start_color='FFFFC7CE', end_color='FFFFC7CE', fill_type='solid')

    # Add summary row
    summary_row = len(data) + 3
    ws.cell(row=summary_row, column=1, value='TOTALS').font = Font(bold=True)

    for col in range(4, 10):  # Numeric columns
        col_letter = get_column_letter(col)
        formula = f'=SUM({col_letter}2:{col_letter}{len(data) + 1})'
        cell = ws.cell(row=summary_row, column=col, value=formula)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    wb.save(output_path)
    return output_path


def create_song_matrix(students_data: list, songs_data: list, output_path: str) -> str:
    """
    Create a song progress matrix with students as rows and songs as columns.

    Args:
        students_data: List of {user, songs} where songs has status per song
        songs_data: List of all songs [{id, title, author}, ...]
        output_path: Where to save the Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Song Matrix"

    # Header row: Student name + all song titles
    ws.cell(row=1, column=1, value='Student').fill = HEADER_FILL
    ws.cell(row=1, column=1).font = HEADER_FONT
    ws.column_dimensions['A'].width = 25

    for col, song in enumerate(songs_data, start=2):
        cell = ws.cell(row=1, column=col, value=song.get('title', 'Unknown'))
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(textRotation=45, horizontal='center')
        ws.column_dimensions[get_column_letter(col)].width = 12

    # Data rows
    for row_num, student in enumerate(students_data, start=2):
        user = student.get('user', {})
        name = f"{user.get('firstName', '')} {user.get('lastName', '')}".strip()
        if not name:
            name = user.get('email', 'Unknown').split('@')[0]

        ws.cell(row=row_num, column=1, value=name)

        # Map student's song statuses
        student_songs = {s.get('id'): s.get('status', 'to_learn') for s in student.get('songs', [])}

        for col, song in enumerate(songs_data, start=2):
            status = student_songs.get(song.get('id'), '')
            cell = ws.cell(row=row_num, column=col, value=status.replace('_', ' ').title() if status else '')
            cell.alignment = Alignment(horizontal='center')

            if status in STATUS_FILLS:
                cell.fill = STATUS_FILLS[status]

    # Freeze first row and column
    ws.freeze_panes = 'B2'

    wb.save(output_path)
    return output_path


def main():
    parser = argparse.ArgumentParser(description='Generate student roster Excel export')
    parser.add_argument('--data', required=True, help='JSON file with student data')
    parser.add_argument('--output', required=True, help='Output Excel path')
    parser.add_argument('--type', choices=['roster', 'matrix'], default='roster', help='Export type')
    args = parser.parse_args()

    with open(args.data, 'r') as f:
        data = json.load(f)

    if args.type == 'roster':
        output = create_roster_export(data, args.output)
    else:
        # For matrix, expect {students: [...], songs: [...]}
        output = create_song_matrix(data['students'], data['songs'], args.output)

    print(f"Export generated: {output}")


if __name__ == '__main__':
    main()
